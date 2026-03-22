"""
AI Office Pilot - File Analyzer
Read and understand file contents
"""

from pathlib import Path
from typing import Optional


class FileAnalyzer:
    """Extract text content from files"""

    MAX_TEXT_LENGTH = 5000
    MAX_PDF_PAGES = 10
    MAX_EXCEL_SHEETS = 3
    MAX_EXCEL_ROWS = 50

    IMAGE_EXTENSIONS = frozenset([".jpg", ".jpeg", ".png", ".gif", ".bmp"])
    TEXT_EXTENSIONS = frozenset([".txt", ".md", ".csv", ".log"])
    EXCEL_EXTENSIONS = frozenset([".xlsx", ".xls"])

    def analyze(self, file_info: dict) -> str:
        """Extract text content from a file"""
        path = Path(file_info["path"])
        extension = file_info["extension"]

        try:
            if extension == ".pdf":
                return self._read_pdf(path)
            elif extension in self.TEXT_EXTENSIONS:
                return self._read_text(path)
            elif extension == ".docx":
                return self._read_docx(path)
            elif extension in self.EXCEL_EXTENSIONS:
                return self._read_excel(path)
            elif extension in self.IMAGE_EXTENSIONS:
                return f"[Image file: {file_info['name']}]"
            else:
                return self._read_text(path)
        except Exception as e:
            return f"[Could not read file: {e}]"

    def _read_pdf(self, path: Path) -> str:
        """Extract text from PDF"""
        try:
            from PyPDF2 import PdfReader

            reader = PdfReader(str(path))
            text = ""
            for page in reader.pages[: self.MAX_PDF_PAGES]:
                text += page.extract_text() or ""
            return text[: self.MAX_TEXT_LENGTH]
        except ImportError:
            return "[PDF reader not installed]"

    def _read_text(self, path: Path) -> str:
        """Read plain text file"""
        try:
            return path.read_text(errors="replace")[: self.MAX_TEXT_LENGTH]
        except OSError:
            return "[Could not read text file]"

    def _read_docx(self, path: Path) -> str:
        """Read Word document"""
        try:
            from docx import Document

            doc = Document(str(path))
            text = "\n".join([p.text for p in doc.paragraphs])
            return text[: self.MAX_TEXT_LENGTH]
        except ImportError:
            return "[DOCX reader not installed]"

    def _read_excel(self, path: Path) -> str:
        """Read Excel spreadsheet"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True)
            text = ""
            for sheet in wb.sheetnames[: self.MAX_EXCEL_SHEETS]:
                ws = wb[sheet]
                for row in ws.iter_rows(max_row=self.MAX_EXCEL_ROWS, values_only=True):
                    text += " | ".join(str(c) for c in row if c) + "\n"
            return text[: self.MAX_TEXT_LENGTH]
        except ImportError:
            return "[Excel reader not installed]"
