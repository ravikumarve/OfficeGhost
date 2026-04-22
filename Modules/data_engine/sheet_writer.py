"""
AI Office Pilot - Spreadsheet Writer
Write data to local Excel or Google Sheets
"""

from pathlib import Path
from datetime import datetime
from typing import Optional

from core.config import Config


class SheetWriter:
    """Write extracted data to spreadsheets"""

    HEADERS = ["Date", "Vendor", "Amount", "Currency", "Category", "Invoice #", "Processed At"]
    CSV_HEADERS = "Date,Vendor,Amount,Currency,Category,Invoice#,ProcessedAt\n"
    DEFAULT_CURRENCY = "USD"
    DATE_FORMAT = "%Y-%m-%d %H:%M"

    def __init__(self) -> None:
        self.spreadsheet_path: Path = Config.LOCAL_SPREADSHEET

    def ensure_spreadsheet(self) -> None:
        """Create spreadsheet if it doesn't exist"""
        if not self.spreadsheet_path.exists():
            self._create_spreadsheet()

    def write_row(self, data: dict) -> Optional[int]:
        """Write a row of data to the spreadsheet"""
        self.ensure_spreadsheet()

        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(self.spreadsheet_path))
            ws = wb.active

            next_row = ws.max_row + 1

            ws.cell(row=next_row, column=1, value=data.get("date", ""))
            ws.cell(row=next_row, column=2, value=data.get("vendor", ""))
            ws.cell(row=next_row, column=3, value=data.get("amount", 0))
            ws.cell(row=next_row, column=4, value=data.get("currency", self.DEFAULT_CURRENCY))
            ws.cell(row=next_row, column=5, value=data.get("category", ""))
            ws.cell(row=next_row, column=6, value=data.get("invoice_number", ""))
            ws.cell(row=next_row, column=7, value=datetime.now().strftime(self.DATE_FORMAT))

            wb.save(str(self.spreadsheet_path))
            return next_row

        except ImportError:
            return self._write_csv(data)

    def _create_spreadsheet(self) -> None:
        """Create new spreadsheet with headers"""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"

            for col, header in enumerate(self.HEADERS, 1):
                ws.cell(row=1, column=col, value=header)

            self.spreadsheet_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(self.spreadsheet_path))

        except ImportError:
            self.spreadsheet_path = self.spreadsheet_path.with_suffix(".csv")
            self.spreadsheet_path.parent.mkdir(parents=True, exist_ok=True)
            self.spreadsheet_path.write_text(self.CSV_HEADERS)

    def _write_csv(self, data: dict) -> Optional[int]:
        """CSV fallback writer"""
        csv_path = self.spreadsheet_path.with_suffix(".csv")
        with open(csv_path, "a") as f:
            row = (
                f"{data.get('date', '')},"
                f"{data.get('vendor', '')},"
                f"{data.get('amount', 0)},"
                f"{data.get('currency', self.DEFAULT_CURRENCY)},"
                f"{data.get('category', '')},"
                f"{data.get('invoice_number', '')},"
                f"{datetime.now().strftime(self.DATE_FORMAT)}\n"
            )
            f.write(row)
        return None
